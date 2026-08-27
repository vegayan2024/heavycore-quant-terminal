# -*- coding: utf-8 -*-
import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

sources_dir = r"C:\个人文件夹\github\terminal\docs\data\sources"
excel_files = [f for f in os.listdir(sources_dir) if f.startswith("xyzq") and f.endswith(".xlsx")]
print("Found XYZQ files:", excel_files)

target_file = os.path.join(sources_dir, sorted(excel_files)[-1])
print(f"Targeting: {target_file}")

# read_only=True is extremely fast for huge Excel files
wb = openpyxl.load_workbook(target_file, read_only=True)
print("\n=== Sheets in Workbook ===")
for idx, name in enumerate(wb.sheetnames):
    print(f"[{idx}] {name}")

# Let's inspect key sheets: e.g. 价格 价差 指标总表, 装置开工, 原油, 氯碱, 农药, 化肥, etc.
key_sheets = [s for s in wb.sheetnames if any(k in s for k in ["价格", "价差", "开工", "库存", "化肥", "氯碱", "涤纶", "聚氨酯", "磷", "农药", "C1", "C2"])]
print("\n=== Key Relevant Sheets ===", key_sheets)

for sname in key_sheets[:6]:
    ws = wb[sname]
    print(f"\n------------------------------------------------------------")
    print(f"Sheet: {sname}")
    print(f"------------------------------------------------------------")
    count = 0
    for row in ws.iter_rows(values_only=True):
        non_empty = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
        if non_empty:
            print(f"Row {count+1}: {non_empty[:8]}")
            count += 1
        if count >= 8:
            break
