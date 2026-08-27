# -*- coding: utf-8 -*-
"""
从兴业证券化工数据库中精确提取各品种历史时序与最新开工率数据
"""
import os
import sys
import json
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

sources_dir = r"C:\个人文件夹\github\terminal\docs\data\sources"
target_file = os.path.join(sources_dir, "xyzq-chemical-prices-20260705.xlsx")
wb = openpyxl.load_workbook(target_file, data_only=True)

# 1. 提取 化肥 表时序 (尿素、价差)
ws_fert = wb["化肥"]
dates = []
urea_prices = []
urea_spread_coal = []
for r in ws_fert.iter_rows(values_only=True, min_row=3, max_col=5):
    if r[0] is not None and str(r[0]).startswith("202"):
        dt_str = str(r[0])[:10]
        p = float(r[1]) if r[1] is not None and str(r[1]).replace('.','').isdigit() else 0
        s = float(r[2]) if r[2] is not None and str(r[2]).replace('-','').replace('.','').isdigit() else 0
        dates.append(dt_str)
        urea_prices.append(p)
        urea_spread_coal.append(s)

# 2. 提取 氯碱 表时序 (纯碱、PVC、烧碱)
ws_cl = wb["氯碱"]
cl_dates = []
soda_ash_prices = [] # 纯碱
soda_ash_spread = []
pvc_prices = []
pvc_spread = []
for r in ws_cl.iter_rows(values_only=True, min_row=3, max_col=10):
    if r[0] is not None and str(r[0]).startswith("202"):
        dt_str = str(r[0])[:10]
        cl_dates.append(dt_str)
        # Check columns
        p_soda = float(r[1]) if len(r)>1 and r[1] is not None and str(r[1]).replace('.','').isdigit() else 0
        s_soda = float(r[2]) if len(r)>2 and r[2] is not None and str(r[2]).replace('-','').replace('.','').isdigit() else 0
        soda_ash_prices.append(p_soda)
        soda_ash_spread.append(s_soda)

# 3. 提取 农药 表时序 (草甘膦、双甘膦、甘氨酸)
ws_pest = wb["农药"]
pest_dates = []
glyphosate_prices = []
for r in ws_pest.iter_rows(values_only=True, min_row=3, max_col=8):
    if r[0] is not None and str(r[0]).startswith("202"):
        pest_dates.append(str(r[0])[:10])
        p_gly = float(r[6]) if len(r)>6 and r[6] is not None and str(r[6]).replace('.','').isdigit() else 0
        glyphosate_prices.append(p_gly)

# 4. 提取 开工率 表
ws_op = wb["开工率"]
operating_rates = {}
for r in ws_op.iter_rows(values_only=True, min_row=2, max_col=6):
    # col 1 is 品种, col 2 is 开工率
    if len(r) >= 3 and r[1] and r[2] is not None:
        try:
            val = float(r[2])
            operating_rates[str(r[1]).strip()] = round(val * 100, 1)
        except:
            pass
    if len(r) >= 6 and r[4] and r[5] is not None:
        try:
            val = float(r[5])
            operating_rates[str(r[4]).strip()] = round(val * 100, 1)
        except:
            pass

print("=== 尿素数据 (最近 8 期) ===")
print("Dates:", dates[:8])
print("Urea Prices (元/吨):", urea_prices[:8])
print("Urea Spread (元/吨):", urea_spread_coal[:8])

print("\n=== 农药草甘膦数据 (最近 8 期) ===")
print("Glyphosate Prices (元/吨):", glyphosate_prices[:8])

print("\n=== 百川盈孚开工率提取结果 (部分展示) ===")
for k in ["合成氨", "黄磷", "磷矿石", "二铵", "磷酸", "复合肥", "DMF", "环氧丙烷（PO）"]:
    print(f" • {k}: {operating_rates.get(k, 'N/A')}%")

# Save extracted result to json for easy pipeline integration
extracted_data = {
    "urea": {"dates": dates[:16], "prices": urea_prices[:16], "spreads": urea_spread_coal[:16]},
    "soda_ash": {"dates": cl_dates[:16], "prices": soda_ash_prices[:16], "spreads": soda_ash_spread[:16]},
    "glyphosate": {"dates": pest_dates[:16], "prices": glyphosate_prices[:16]},
    "operating_rates": operating_rates
}

with open(os.path.join(os.path.dirname(__file__), "extracted_chemical_data.json"), "w", encoding="utf-8") as f:
    json.dump(extracted_data, f, ensure_ascii=False, indent=2)

print("\n✅ 数据提取完毕并保存至 extracted_chemical_data.json！")
