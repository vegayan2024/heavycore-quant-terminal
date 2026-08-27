# -*- coding: utf-8 -*-
import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

sources_dir = r"C:\个人文件夹\github\terminal\docs\data\sources"
target_file = os.path.join(sources_dir, "xyzq-chemical-prices-20260705.xlsx")

wb = openpyxl.load_workbook(target_file, read_only=True)

print("=== ALL SHEETS IN XYZQ EXCEL ===")
for i, name in enumerate(wb.sheetnames):
    print(f"{i}: {name}")

# Let's inspect '化肥' (Fertilizer) and '氯碱' or 'C1'
target_sheets = [s for s in wb.sheetnames if s in ["化肥", "氯碱", "C1", "C2", "C3", "农药", "价格 价差 库存等指标汇总表"]]
print("\nTargeting sheets for deep inspection:", target_sheets)

for sname in target_sheets:
    ws = wb[sname]
    print(f"\n====================== Sheet: {sname} ======================")
    rows = list(ws.iter_rows(values_only=True, max_row=25, max_col=12))
    for r_idx, row in enumerate(rows):
        cells = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
        if cells:
            print(f"Row {r_idx+1}: {cells[:8]}")
